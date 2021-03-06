#!/usr/bin/env python3

# UPDATE THIS EVERY TIME A NEW RELEASE IS PACKAGED!
VERSION = "2.4.5"

# Spyglass
# Source code by Derps aka Panzer Vier
# Modifications made by Khronion (KH)
# Updated to py3 by 9003

import urllib
import gzip
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import xml.etree.cElementTree as ElementTree
import math
from datetime import datetime
import os
import sys
import urllib.request


logpath = "debug.log"


# method for writing a debug log
def write_log(text):
    with open(logpath, "a") as out:
        print(str(out) + '[{:%Y-%m-%d %H:%M:%S}] '.format(datetime.now()) + text)


# method for getting user input
def query(text, options):
    while True:
        response = input(text)
        if response in options:
            return response

# parse arguments, if any...

# show help message and terminate
if "-h" in sys.argv or "--help" in sys.argv:
    print( "Spyglass {}: Generate NationStates region update timesheets.\n".format(VERSION))
    print( "Developed by Panzer Vier, with additions by Khronion translated to py3 by 9003\n")
    print( "usage: {} [-h] [-n NATION] [-o OUTFILE] [-s | -l PATH]\n".format(sys.argv[0]))
    print( "Optional arguments:\n" \
          " -h           Show this help message and exit.\n" \
          " -n NATION    Specify Nation to identify user by. In order to comply with \n" \
          "              NationStates API rules, this must be the user's nation. Use\n" \
          "              underscores instead of spaces.\n" \
          " -o OUTFILE   File to output the generated timesheet in XLSX format to.\n" \
          " -s           Suppress creating a debug log file. Log files are written to\n" \
          "              the current working directory.\n" \
          " -l PATH      Write debug log to specified path.\n" \
          " -m           Generate a minimized sheet without WFEs and embassies\n")
    print( "If run without arguments, Spyglass runs in interactive mode and outputs to its\n" \
          "working directory.")
    sys.exit()

process_embassies = False
log = True

SpeedOverride = False
MinorTime = 2640
MajorTime = 3540

now = datetime.now()
YMD = '%s-%s-%s' % (now.year, now.month, now.day)

# set nation name
if "-n" in sys.argv:
    UAgent = sys.argv[sys.argv.index("-n") + 1]
else:
    print( "Spyglass {}: Generate NationStates region update timesheets.".format(VERSION))
    UAgent = str(input('Nation Name: '))
    tag = str(input('Tag to search (Default is Fascist)'))
    if not tag:
        tag="Fascist"
    filename = 'SpyglassSheet' + YMD + '.xlsx'

#    if query("Include region embassies? (y/n, defaults to y) ", ['y', 'n', '']) == 'n':
#        process_embassies = False

    # Update lengths are now set to 44m and 59m, per word of [v]
    if query("Do you want to manually specify update lengths? (y/n, defaults to n) ", ['y', 'n', '']) == 'y':
        try:
            MinorTime = int(input("Minor Time, seconds (2640): "))
        except SyntaxError:
            MinorTime = 2640
        try:
            MajorTime = int(input("Major Time, seconds (3540): "))
        except SyntaxError:
            MajorTime = 3540
        SpeedOverride = True

# set output filename
if "-o" in sys.argv:
    filename = sys.argv[sys.argv.index("-o") + 1]
else:
    filename = 'SpyglassSheet' + YMD + '.xlsx'

# enable debug log
if "-s" in sys.argv:
    log = False

if "-m" in sys.argv:
    process_embassies = False

else:
    if "-l" in sys.argv:
        logpath = sys.argv[sys.argv.index("-l") + 1]
    write_log("INFO Spyglass started with arguments: " + " ".join(sys.argv[1:]))
    write_log("INFO User Nation: " + UAgent)
    write_log("INFO Out File: " + filename)

# Set headers as required by NS TOS
headers = {
    'User-Agent': 'Spyglass. Currently in use by {} (Authenticating). '
                  'Source code: https://github.com/khronion/Spyglass'.format(UAgent)}

# Verify specified nation is valid -- terminate if not
try:
    testReq = urllib.request.Request(
        'https://www.nationstates.net/cgi-bin/api.cgi?nation=' + UAgent.replace(' ', '_') + "&q=influence", None,
        headers)
    testhtml = urllib.request.urlopen(testReq).read()
    headers = {
        'User-Agent': 'Spyglass. Currently in use by ' + UAgent + '. Source code: https://github.com/khronion/Spyglass'}
except urllib.error.HTTPError as e:
    print( "Nation not found. Be sure to input the name of a nation that actually exists.")
    if log:
        write_log("ERR  {} is not a valid nation. Terminating.".format(UAgent))
    sys.exit()

if log:
    write_log("INFO Minor length: " + str(MinorTime))
    write_log("INFO Major length: " + str(MajorTime))


# Pulling a list of regions that are founderless and non-passworded. Eventually, we'll go through and highlight those
# on the sheet

if log:
    write_log("INFO Downloading data dump...")

# Total number of queries is low, so rate limit is unnecessary
req = urllib.request.Request('https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=-password', None, headers)
req2 = urllib.request.Request('https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=founderless', None, headers)
fashBash = urllib.request.Request('https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=-password,'+tag, None, headers)
peakFashBash = urllib.request.Request('https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=-password,'+tag+',founderless', None, headers)
html = urllib.request.urlopen(req).read()
html2 = urllib.request.urlopen(req2).read()
html3 = urllib.request.urlopen(fashBash).read()
html4 = urllib.request.urlopen(peakFashBash).read()

# Grabbing the data dump and saving
print( "Pulling Data Dump...")
urllib.request.urlretrieve('https://www.nationstates.net/pages/regions.xml.gz', 'regions.xml.gz')
if log:
    write_log("INFO Download complete!")

redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
greenFill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')
yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
blueFill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
pinkFill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
#
# Un-gzipping
# TODO: instead of saving and reading from disk, stream regions.xml.gz directly to the parser
with gzip.open('regions.xml.gz', 'rb') as infile:
    with open('regions.xml', 'w') as outfile:
        for line in infile:
            outfile.write(str(line))

# Opening up my virtual sheet. Gotta find a better name for it, sometime. The pink tab colour's pretty sweet, tho.
wb = Workbook()
ws = wb.active
ws.title = "Spyglass Timesheet"
ws.sheet_properties.tabColor = "FFB1B1"

RegionList = []
RegionURLList = []
RegionWFEList = []
RegionEmbassyList = []
NumNationList = []
DelVoteList = []
ExecList = []
MajorList = []

# Sanitize our founderless regions list a wee bit, 'cause at the moment, it's xml, and xml is gross.
print( "Processing data...")
UnfoundedList = ElementTree.fromstring(html2).find('REGIONS').text.split(',')
PWlessList = ElementTree.fromstring(html).find('REGIONS').text.split(',')
FashestOfBashes =  ElementTree.fromstring(html3).find('REGIONS').text.split(',')
FashestOfBashesOfAll =  ElementTree.fromstring(html4).find('REGIONS').text.split(',')


# TODO: instead of saving and reading from disk, stream regions.xml.gz directly to the parser
with open('regions.xml', 'r') as myfile:
    regions = myfile.read()

# Pulling, in order, region names, converting to a region url, number of nations in that region, and voting power that
# delegate has.

# KH: Replaced some assignments with augmented assignments (e.g. x += y instead of x = x + y)
#input(regions)
root = regions.split("\\n\'b\'")
for x in root:
    if '<NAME>' in x and '</NAME>' in x:
        x=x.replace("</NAME>","")
        x=x.replace("<NAME>","")
        RegionList.append(x)
        UrlString = '=HYPERLINK("https://www.nationstates.net/region=' + x + '")'
        UrlString=UrlString.replace(' ', '_')
        RegionURLList.append(UrlString)
    if '<NUMNATIONS>' in x and '</NUMNATIONS>' in x:
        x=x.replace("</NUMNATIONS>","")
        x=x.replace("<NUMNATIONS>","")
        NumNationList.append(x)
    if '<DELEGATEVOTES>' in x and '</DELEGATEVOTES>' in x:
        x=x.replace("</DELEGATEVOTES>","")
        x=x.replace("<DELEGATEVOTES>","")
        DelVoteList.append(x)
    if '<DELEGATEAUTH>' in x and '</DELEGATEAUTH>' in x:
        x=x.replace("</DELEGATEAUTH>","")
        x=x.replace("<DELEGATEAUTH>","")
        if 'X' in x:
            ExecList.append(True)
        else:
            ExecList.append(False)
    # KH: pull major times from daily dump
    if '<LASTUPDATE>' in x and '</LASTUPDATE>' in x:
        x=x.replace("</LASTUPDATE>","")
        x=x.replace("<LASTUPDATE>","")
        MajorList.append(x)
    # KH: gather WFE info
    if '<FACTBOOK>' in x and '</FACTBOOK>' in x:
        x=x.replace("</FACTBOOK>","")
        x=x.replace("<FACTBOOK>","")
        try:
            if ['=', '+', "-", "@"] in x:
                x="'"+x
            RegionWFEList.append(x)
        except TypeError: # no WFE
            RegionWFEList.append(" ")
    # KH: gather embassy list
    if '<EMBASSY>' in x and '</EMBASSY>' in x:
        x=x.replace("</EMBASSY>","")
        x=x.replace("<EMBASSY>","")
        if process_embassies:
            RegionEmbassyList.append(x)
# Grabbing the cumulative number of nations that've updated by the time a region has.
# The first entry is zero because time calculations need to reflect the start of region update, not the end
CumulNationList = [0]
for a in NumNationList:
    if len(CumulNationList) == 0:
        CumulNationList.extend([int(a)])
    else:
        CumulNationList.extend([int(a) + CumulNationList[-1]])

# Calculate speed based on total population
CumulNations = CumulNationList[-1]
MinorNatTime = float(MinorTime) / CumulNations
MajorNatTime = float(MajorTime) / CumulNations
MinTime = []
MajTime = []

# Getting the approximate major/minor update times.
for a in CumulNationList:
    temptime = int(a * MinorNatTime)
    tempsecs = temptime % 60
    tempmins = int(math.floor(temptime / 60) % 60)
    temphours = int(math.floor(temptime / 3600))
    MinTime.extend(['%s:%s:%s' % (temphours, tempmins, tempsecs)])

# If user specifies update length, use special handling.
if SpeedOverride:
    for a in CumulNationList:
        temptime = int(a) * int(MajorNatTime)
        tempsecs = temptime % 60
        tempmins = int(math.floor(temptime / 60) % 60)
        temphours = int(math.floor(temptime / 3600))
        MajTime.extend(['%s:%s:%s' % (temphours, tempmins, tempsecs)])
else:
    for a in MajorList:
        temptime = int(a) - int(MajorList[0])
        tempsecs = temptime % 60
        tempmins = int(math.floor(temptime / 60) % 60)
        temphours = int(math.floor(temptime / 3600))
        MajTime.extend(['%s:%s:%s' % (temphours, tempmins, tempsecs)])

# Splashing some headers and stuff onto the spreadsheet for legibility purposes!
ws['A1'].value = 'Regions'
ws['B1'].value = 'Region Link'
ws['C1'].value = '# Nations'
ws['D1'].value = 'Tot. Nations'
ws['E1'].value = 'Minor Upd. (est)'
ws['F1'].value = 'Major Upd. (true)'
ws['G1'].value = 'Del. Votes'
ws['H1'].value = 'Del. Endos'
if process_embassies:
    ws['I1'].value = 'Embassies'
ws['J1'].value = 'WFE'

ws['L1'].value = 'World '
ws['M1'].value = 'Data'
ws['L2'].value = 'Nations'
ws['L3'].value = 'Last Major'
ws['L4'].value = 'Secs/Nation'
ws['L5'].value = 'Nations/Sec'
ws['O2'].value = 'Green'
ws['P2'].value = 'Founderless/passwordless/exec delegate ~'
ws.cell(row=2, column=15).fill = greenFill
ws.cell(row=2, column=16).fill = greenFill
ws['O3'].value = 'yellow'
ws['P3'].value = 'Passwordless/exec delegate ~'
ws.cell(row=3, column=15).fill = yellowFill
ws.cell(row=3, column=16).fill = yellowFill
ws['O4'].value = 'Red'
ws['P4'].value = 'Passworded*'
ws.cell(row=4, column=15).fill = redFill
ws.cell(row=4, column=16).fill = redFill
ws['O5'].value = 'Blue'
ws['P5'].value = tag + ' w/ founder and no password !'
ws.cell(row=5, column=15).fill = blueFill
ws.cell(row=5, column=16).fill = blueFill
ws['O6'].value = 'Pink'
ws['P6'].value = tag+' W/O founder and no password !!'
ws.cell(row=6, column=15).fill = pinkFill
ws.cell(row=6, column=16).fill = pinkFill
ws['L6'].value = 'Last Minor'
ws['L7'].value = 'Secs/Nation'
ws['L8'].value = 'Nations/Sec'
ws['L10'].value = 'Spyglass Version'
ws['L11'].value = 'Date Generated'
ws['M2'].value = CumulNations
ws['M3'].value = int(MajorList[-1]) - int(MajorList[0])
ws['M4'].value = float(MajorList[-1]) - float(MajorList[0]) / float(CumulNations)
ws['M5'].value = 1 / float(MajorList[-1]) - float(MajorList[0]) / float(CumulNations)
ws['M6'].value = MinorTime
ws['M7'].value = MinorNatTime
ws['M8'].value = 1 / MinorNatTime
ws['M10'].value = VERSION
ws['M11'].value = YMD

# There's probably a better way of doing this, but my coding skills are dubious :^)
# Anyways, actually pasting the information from our various lists into the spreadsheet.
counter = 0
for a in RegionList:
    # If a region's founderless, highlight it for easy reference. Add a tilde, 'cause my spreadsheet program doesn't
    # do filtering by colour

    # TODO: document specific key characters and colors that can be used to sort
    b = a
    # KH: ~ indicates hittable
    # KH: yellow = passwordless and exec delegate
    if a in PWlessList and ExecList[counter] is True:
        ws.cell(row=counter + 2, column=1).fill = yellowFill
        ws.cell(row=counter + 2, column=2).fill = yellowFill
        b = a + '~'
    # KH: green = founderless and passwordless
    if a in UnfoundedList and a in PWlessList:
        ws.cell(row=counter + 2, column=1).fill = greenFill
        ws.cell(row=counter + 2, column=2).fill = greenFill
        b = a + '~'
    # KH: red = passwordless
    if a not in PWlessList:
        ws.cell(row=counter + 2, column=1).fill = redFill
        ws.cell(row=counter + 2, column=2).fill = redFill
        b = a + '*'
    if a in FashestOfBashes:
        ws.cell(row=counter + 2, column=1).fill = blueFill
        ws.cell(row=counter + 2, column=2).fill = blueFill
        b = a + ' !'
    if a in FashestOfBashesOfAll:
        ws.cell(row=counter + 2, column=1).fill = pinkFill
        ws.cell(row=counter + 2, column=2).fill = pinkFill
        b = a + ' !!'

    ws.cell(row=counter + 2, column=1).value = b
    ws.cell(row=counter + 2, column=2).value = RegionURLList[counter]
    ws.cell(row=counter + 2, column=3).value = NumNationList[counter]
    ws.cell(row=counter + 2, column=4).value = CumulNationList[counter]
    ws.cell(row=counter + 2, column=5).value = MinTime[counter]
    ws.cell(row=counter + 2, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=6).value = MajTime[counter]
    ws.cell(row=counter + 2, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=7).value = DelVoteList[counter]
    ws.cell(row=counter + 2, column=8).value = float(DelVoteList[counter]) - 1
    #ws.cell(row=counter + 2, column=9).value = RegionEmbassyList[counter]
    #ws.cell(row=counter + 2, column=10).value = RegionWFEList[counter]
    ws.cell(row=counter + 2, column=11).value = " "

    # Highlight delegate-less regions. They're good for tagging, or whatever~
    if DelVoteList[counter] == 0:
        ws.cell(row=counter + 2, column=8).fill = redFill
    counter += 1

# You know those situations where you can't quite get code to work, and kinda fumble around until you find something
# that does?
#
# I'm 90% sure this isn't the way to do it, but I couldn't get it working otherwise.
# Anyways, setting the region name column's width, so that it doesn't cut everything off.
sheet = wb['Spyglass Timesheet']
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['P'].width = 45
sheet['J1'].alignment = Alignment(horizontal="right")

if log:
    write_log("INFO Done processing data! Saving sheet.")

# Really should just name the sheets 'Derps is amazing in every conceivable way'. Would be some free ego-massage.
print( "Saving Sheet...")
wb.save(filename)
# Deleting the dump, 'cause not needed anymore. Let's keep things vaguely clean, neh?

if log:
    write_log("INFO Successfully saved to " + filename)

print( "Cleaning up...")
os.remove('regions.xml.gz')
os.remove('regions.xml')

if log:
    write_log("INFO Spyglass run complete. Exiting...")

sys.exit()
